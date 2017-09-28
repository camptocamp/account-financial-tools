# -*- coding: utf-8 -*-
# Copyright 2017 Camptocamp SA
# License LGPL-3.0 or later (http://www.gnu.org/licenses/lgpl.html).

from __future__ import division
from datetime import datetime
from odoo import api, fields, models, _
from odoo.exceptions import Warning
from odoo.modules.module import get_module_path
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles.borders import Border, Side


class IntrastatDoc(models.TransientModel):
    """
        Creates two types of intrastat document from a stock picking,
        a packing list and a commercial invoice. The documents are based on
        Excel file templates.
    """
    _name = 'intrastat.doc'
    picking_id = fields.Integer()
    document_type = fields.Char()
    total = fields.Float(default=0)
    unity_of_measure = fields.Char()

    stock_pick = None
    sale_order = None

    MODULE_PATH = get_module_path('intrastat_doc')
    INVOICE_TMPL = MODULE_PATH + '/templates/commercial_invoice_tmpl.xlsx'
    PACKING_LIST_TMPL = MODULE_PATH + '/templates/packing_list_tmpl.xlsx'
    TOP_LOGO = MODULE_PATH + '/templates/logo.png'

    cell_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    @api.multi
    def download(self):
        """
            Create the Excel document and returns it
        """
        self.ensure_one()
        self.stock_pick = self.env['stock.picking'].browse(
                self.picking_id).exists()
        if not self.stock_pick:
            raise Warning(_('The stock picking could not be found.'))
        self.sale_order = self.env['sale.order'].search(
                [('procurement_group_id', '=', self.stock_pick.group_id.id)])
        # Set the unity of measure depending on the type of document
        if self.document_type == 'packinglist':
            wb = load_workbook(self.PACKING_LIST_TMPL)
            self.unity_of_measure = 'kg'
        else:
            # For the commercial invoice use the currency of the sale order
            wb = load_workbook(self.INVOICE_TMPL)
            self.unity_of_measure = (self.sale_order.holding_currency_id.name
                                     or '')
        ws = wb.active
        self._set_header(ws)
        current_line = 18
        for line in self._get_detail_lines():
            self._add_line(ws, current_line, line, True)
            if len(line) > 3:
                self.total += line[3]
            current_line += 1
        self._set_footer(ws, current_line)
        img = Image(self.TOP_LOGO)
        ws.add_image(img, 'A1')
        return save_virtual_workbook(wb)

    def _add_line(self, ws, row_index, data, with_border):
        for index, value in enumerate(data):
            col_index = chr(ord('D') + index)
            ws[col_index + str(row_index)] = value
        if with_border:
            for col_index in range(4, 8):
                ws.cell(row=row_index,
                        column=col_index).border = self.cell_border

    @api.multi
    def _set_header(self, ws):
        """ Set the header part of the excel document """
        self.ensure_one()
        # Ship to address and contact
        ship_to_address = self.sale_order.partner_shipping_id.contact_address
        self._insert_address(ws, 4, 'D', ship_to_address)
        ws['D10'] = self.stock_pick.partner_id.name or ''
        ws['D11'] = self.stock_pick.partner_id.phone or ''
        if self.document_type == 'packinglist':
            # Warehouse address and contact
            warehouse = self.stock_pick.picking_type_id.warehouse_id
            warehouse_address = warehouse.partner_id.contact_address or ''
            self._insert_address(ws, 4, 'G', warehouse_address)
            warehouse_contact = warehouse.partner_id.child_ids.filtered(
                    lambda r: r.type == 'contact')
            if warehouse_contact:
                ws['G10'] = warehouse_contact.name or ''
                ws['G11'] = warehouse_contact.phone or ''
            else:
                ws['G10'] = warehouse.partner_id.name or ''
                ws['G11'] = warehouse.partner_id.phone or ''
        else:
            bill_address = self.sale_order.partner_invoice_id.contact_address
            self._insert_address(ws, 4, 'G', bill_address)
        # Header details
        today = datetime.now()
        ws['D15'] = today.strftime('%y%m%d')
        ws['E15'] = today.strftime('%m/%d/%y')
        if self.sale_order:
            ws['F15'] = self.sale_order.name
            ws['G15'] = self.sale_order.project_id.name or ''

    @api.multi
    def _set_footer(self, ws, row):
        self.ensure_one()
        row += 1
        self._add_line(ws, row, ['', '', 'Packed in XXX'], True)
        self._add_line(ws, row, ['', '', 'Size: L*|*H'], True)
        self._add_line(ws, row, ['', '', 'INCOTERM'], True)
        self._add_line(ws, row,
                       ['', '', 'Total {}'.format(self.unity_of_measure),
                        self.total], True)

    @api.multi
    def _insert_address(self, ws, row, column, address):
        self.ensure_one()
        if not address:
            return
        for line in address.split('\n'):
            if len(line):
                ws[column + str(row)] = line
                row += 1

    @api.multi
    def _lines_product(self, product, qty, serial, value, no_hs=None):
        """ Constructs the detail lines for one product """
        self.ensure_one()
        lines = []
        if self.document_type == 'packinglist':
            column4 = product.weight * qty
        else:
            column4 = value
        lines.append([qty, product.uom_po_id.name, product.name, column4])
        if serial:
            lines.append(['', '', u'Serial # {}'.format(serial)])
        hs_code = product.hs_code or ''
        if no_hs:
            hs_code = 'MISSING HS CODE'
        lines.append(['', '', u'HS Code: {}'.format(hs_code)])

        lines.append([''])
        lines.append(['', '', u'Country of Origin: {}'.format(
            product.origin_country_id.name or '')])
        lines.append([''])
        return lines

    @api.multi
    def _get_detail_lines(self):
        """ Construct all the detail lines for the document """
        self.ensure_one()
        lines = []
        for move in self.stock_pick.move_lines:
            product_moved = move.product_id
            price_total = move.procurement_id.sale_line_id.price_total or 0
            quants = move.quant_ids | move.reserved_quant_ids
            qty_moved = sum(quants.mapped('qty'))
            if product_moved.hs_code:
                # Dealing with a simple product
                serial = ' / '.join(quants.mapped('lot_id'))
                lines.extend(self._lines_product(
                    product_moved, qty_moved, serial, price_total))
            else:
                # Dealing with a product created from other products
                # Using the production order to get total qty produced
                production_orders = self.env['mrp.production'].browse()
                for q in quants:
                    production_orders |= q.history_ids.mapped('production_id')
                if not production_orders:
                    # Dealing with a simple product without hs_code
                    serial = ' / '.join(quants.mapped('lot_id'))
                    lines.extend(self._lines_product(
                        product_moved, qty_moved, serial, price_total,
                        no_hs=True))
                    continue
                production_qty = production_orders[0].product_qty
                # Find all products use for production of the one
                consumed_quants = quants.consumed_quant_ids
                all_products = consumed_quants.mapped('product_id')
                # Spread the value on all product evenly
                value = 0
                if price_total > 0:
                    value = price_total / len(all_products)
                for p in all_products:
                    q = consumed_quants.filtered(
                            lambda r: r.product_id.id == p.id)
                    qty = sum(q.mapped('qty')) * qty_moved / production_qty
                    serial = ' / '.join(q.mapped('lot_id'))
                    lines.extend(self._lines_product(p, qty, serial, value))
        return lines
